 # app_taller/excel_utils.py
import io
from typing import List, Tuple
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .etl_utils import (
    normalizar_rut,
    normalizar_nombre,
    normalizar_email,
    normalizar_patente,
    validar_patente,
)
from .models import (
    Usuario,
    Vehiculo,
    Taller,
    OrdenTrabajo,
    SolicitudRepuesto,
    Repuesto,
    MovimientoRepuesto,
)


# =======================
# HELPERS
# =======================

def _safe_dt(dt):
    """Convierte datetimes con tz en naive para Excel."""
    if not dt:
        return None
    if hasattr(dt, "tzinfo") and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt


def _response_from_wb(wb: Workbook, filename: str) -> HttpResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# =======================
# EXPORTAR A EXCEL
# =======================

def exportar_usuarios_xlsx() -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers = ["RUT", "Nombre completo", "Email", "Rol", "Taller", "Teléfono", "Activo"]
    ws.append(headers)

    qs = Usuario.objects.select_related("taller").all().order_by("rut")

    for u in qs:
        ws.append(
            [
                u.rut or "",
                u.nombre_completo or "",
                u.email or "",
                u.rol or "",
                u.taller.nombre if getattr(u, "taller", None) else "",
                u.telefono or "",
                "SI" if u.activo else "NO",
            ]
        )

    return _response_from_wb(wb, "usuarios.xlsx")


def exportar_vehiculos_xlsx() -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehículos"

    headers = [
        "Patente",
        "Marca",
        "Modelo",
        "Año modelo",
        "VIN",
        "Estado",
        "Conductor RUT",
        "Conductor nombre",
        "Fecha creación",
    ]
    ws.append(headers)

    qs = Vehiculo.objects.select_related("conductor_actual").all().order_by("patente")

    for v in qs:
        ws.append(
            [
                v.patente or "",
                v.marca or "",
                v.modelo or "",
                v.año_modelo or "",
                v.vin or "",
                v.estado or "",
                v.conductor_actual.rut if v.conductor_actual else "",
                v.conductor_actual.nombre_completo if v.conductor_actual else "",
                timezone.localtime(v.fecha_creacion).strftime("%Y-%m-%d %H:%M")
                if v.fecha_creacion
                else "",
            ]
        )

    return _response_from_wb(wb, "vehiculos.xlsx")


def exportar_ots_xlsx() -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "OrdenesTrabajo"

    headers = [
        "Número OT",
        "Patente",
        "Taller código",
        "Taller nombre",
        "Estado",
        "Prioridad",
        "Emergencia",
        "Descripción problema",
        "Diagnóstico inicial",
        "Fecha apertura",
        "Fecha finalización",
        "Fecha cierre",
        "Total repuestos",
        "Total mano_obra",
        "Total OT",
    ]
    ws.append(headers)

    qs = (
        OrdenTrabajo.objects.select_related("vehiculo", "taller")
        .all()
        .order_by("-fecha_apertura")
    )

    for ot in qs:
        fecha_ap = _safe_dt(ot.fecha_apertura)
        fecha_fin = _safe_dt(ot.fecha_finalizacion)
        fecha_cierre = _safe_dt(ot.fecha_cierre)

        estado = ot.estado
        if hasattr(ot, "get_estado_display"):
            estado = ot.get_estado_display()

        prioridad = ot.prioridad
        if hasattr(ot, "get_prioridad_display"):
            prioridad = ot.get_prioridad_display()

        ws.append(
            [
                ot.numero_ot,
                ot.vehiculo.patente if ot.vehiculo else "",
                ot.taller.codigo if ot.taller else "",
                ot.taller.nombre if ot.taller else "",
                estado,
                prioridad,
                "SI" if ot.emergencia else "NO",
                ot.descripcion_problema or "",
                ot.diagnostico_inicial or "",
                fecha_ap,
                fecha_fin,
                fecha_cierre,
                float(ot.total_repuestos or 0),
                float(ot.total_mano_obra or 0),
                float(ot.total_ot or 0),
            ]
        )

    return _response_from_wb(wb, "ordenes_trabajo.xlsx")


def exportar_repuestos_xlsx() -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Repuestos"

    headers = [
        "ID",
        "SKU",
        "Nombre",
        "Descripción",
        "Unidad",
        "Precio costo",
        "Categoría",
        "Stock actual",
        "Activo",
        "Fecha creación",
    ]
    ws.append(headers)

    qs = Repuesto.objects.all().order_by("nombre")

    for r in qs:
        ws.append(
            [
                r.id,
                r.sku or "",
                r.nombre or "",
                r.descripcion or "",
                r.unidad or "",
                float(r.precio_costo or 0),
                r.categoria or "",
                r.stock_actual or 0,
                "SI" if r.activo else "NO",
                timezone.localtime(r.fecha_creacion).strftime("%Y-%m-%d %H:%M")
                if r.fecha_creacion
                else "",
            ]
        )

    return _response_from_wb(wb, "repuestos.xlsx")


def exportar_solicitudes_xlsx() -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes Repuestos"

    headers = [
        "ID",
        "N° OT",
        "Patente",
        "SKU repuesto",
        "Nombre repuesto",
        "Cantidad solicitada",
        "Urgente",
        "Estado",
        "N° OC",
        "Proveedor",
        "Fecha entrega estimada",
        "Creado por",
        "Fecha creación",
        "Fecha actualización",
    ]
    ws.append(headers)

    qs = (
        SolicitudRepuesto.objects.select_related(
            "orden_trabajo",
            "orden_trabajo__vehiculo",
            "repuesto",
            "creado_por",
        )
        .all()
        .order_by("-fecha_creacion")
    )

    for s in qs:
        ot = s.orden_trabajo
        veh = getattr(ot, "vehiculo", None)

        ws.append(
            [
                s.id,
                getattr(ot, "numero_ot", "") if ot else "",
                getattr(veh, "patente", "") if veh else "",
                s.repuesto.sku if s.repuesto else "",
                s.repuesto.nombre if s.repuesto else "",
                getattr(s, "cantidad_solicitada", None),
                "Sí" if s.urgente else "No",
                s.estado,
                s.numero_oc or "",
                s.nombre_proveedor or "",
                _safe_dt(s.fecha_entrega_estimada),
                s.creado_por.nombre_completo if s.creado_por else "",
                _safe_dt(s.fecha_creacion),
                _safe_dt(s.fecha_actualizacion),
            ]
        )

    filename = f"solicitudes_repuestos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _response_from_wb(wb, filename)


def exportar_movimientos_xlsx() -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos Repuestos"

    headers = [
        "ID",
        "Fecha movimiento",
        "Taller",
        "N° OT",
        "Patente",
        "SKU repuesto",
        "Nombre repuesto",
        "Tipo movimiento",
        "Cantidad",
        "Costo unitario",
        "Subtotal",
        "Motivo",
        "Movido por",
    ]
    ws.append(headers)

    qs = (
        MovimientoRepuesto.objects.select_related(
            "taller",
            "orden_trabajo",
            "orden_trabajo__vehiculo",
            "repuesto",
            "movido_por",
        )
        .all()
        .order_by("-fecha_movimiento")
    )

    for m in qs:
        ot = m.orden_trabajo
        veh = getattr(ot, "vehiculo", None)
        costo_unit = m.costo_unitario or 0
        cantidad = m.cantidad or 0
        subtotal = costo_unit * cantidad

        ws.append(
            [
                m.id,
                _safe_dt(m.fecha_movimiento),
                m.taller.nombre if m.taller else "",
                getattr(ot, "numero_ot", "") if ot else "",
                getattr(veh, "patente", "") if veh else "",
                m.repuesto.sku if m.repuesto else "",
                m.repuesto.nombre if m.repuesto else "",
                m.tipo_movimiento,
                cantidad,
                float(costo_unit),
                float(subtotal),
                m.motivo or "",
                m.movido_por.nombre_completo if m.movido_por else "",
            ]
        )

    filename = f"movimientos_repuestos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _response_from_wb(wb, filename)


# =======================
# IMPORTAR DESDE EXCEL
# =======================

def importar_usuarios_xlsx(uploaded_file) -> Tuple[int, List[str]]:
    """
    Lee un .xlsx de usuarios y crea/actualiza registros.
    Espera encabezados:
      RUT | Nombre completo | Email | Rol | Taller | Teléfono | Activo
    """
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    errores: List[str] = []
    ok = 0

    # Cabecera
    header = [str(c).strip() if c else "" for c in next(ws.iter_rows(values_only=True))]
    idx = {name: i for i, name in enumerate(header)}

    def _get(row, col, default=""):
        pos = idx.get(col)
        if pos is None or pos >= len(row):
            return default
        val = row[pos]
        return "" if val is None else val

    required = ["RUT", "Nombre completo", "Email"]
    for c in required:
        if c not in idx:
            errores.append(f"Falta columna obligatoria: {c}")
            return 0, errores

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=2):
        rut = normalizar_rut(_get(row, "RUT"))
        nombre = normalizar_nombre(_get(row, "Nombre completo"))
        email = normalizar_email(_get(row, "Email"))
        rol = str(_get(row, "Rol", "")).strip().upper()
        taller_nombre = str(_get(row, "Taller", "")).strip()
        telefono = str(_get(row, "Teléfono", "")).strip()
        activo_txt = str(_get(row, "Activo", "")).strip().upper()

        if not email and not rut:
            # Fila vacía
            continue

        if not email:
            errores.append(f"Fila {row_idx}: Email obligatorio.")
            continue

        activo = True
        if activo_txt in {"NO", "N", "0", "FALSE"}:
            activo = False

        taller_obj = None
        if taller_nombre:
            taller_obj = Taller.objects.filter(nombre__iexact=taller_nombre).first()
            if not taller_obj:
                errores.append(
                    f"Fila {row_idx}: Taller '{taller_nombre}' no encontrado. Se deja sin taller."
                )

        usuario, created = Usuario.objects.get_or_create(
            email=email,
            defaults={
                "rut": rut,
                "nombre_completo": nombre or email,
                "rol": rol or "MECANICO",
                "telefono": telefono,
                "activo": activo,
                "taller": taller_obj,
            },
        )

        if not created:
            if rut:
                usuario.rut = rut
            if nombre:
                usuario.nombre_completo = nombre
            if rol:
                usuario.rol = rol
            if telefono:
                usuario.telefono = telefono
            usuario.activo = activo
            if taller_obj:
                usuario.taller = taller_obj
            usuario.save()

        ok += 1

    return ok, errores


def importar_vehiculos_xlsx(uploaded_file) -> Tuple[int, List[str]]:
    """
    Importa vehículos.
    Encabezados esperados:
      Patente | Marca | Modelo | Año modelo | VIN | Estado
    """
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    errores: List[str] = []
    ok = 0

    # Cabecera
    header = [str(c).strip() if c else "" for c in next(ws.iter_rows(values_only=True))]
    idx = {name: i for i, name in enumerate(header)}

    def _get(row, col, default=""):
        pos = idx.get(col)
        if pos is None or pos >= len(row):
            return default
        val = row[pos]
        return "" if val is None else val

    if "Patente" not in idx:
        errores.append("Falta columna obligatoria: Patente")
        return 0, errores

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=2):
        patente_raw = _get(row, "Patente")
        if not patente_raw:
            continue

        pat = normalizar_patente(patente_raw)
        if not validar_patente(pat):
            errores.append(f"Fila {row_idx}: patente inválida '{patente_raw}'")
            continue

        marca = str(_get(row, "Marca", "")).strip()
        modelo = str(_get(row, "Modelo", "")).strip()
        anio_raw = _get(row, "Año modelo", "")
        vin = str(_get(row, "VIN", "")).strip()
        estado = str(_get(row, "Estado", "")).strip().upper()

        anio_modelo = None
        if anio_raw not in ("", None):
            try:
                anio_modelo = int(anio_raw)
            except Exception:
                errores.append(
                    f"Fila {row_idx}: Año modelo '{anio_raw}' inválido. Se deja en blanco."
                )

        veh, created = Vehiculo.objects.get_or_create(
            patente=pat,
            defaults={
                "marca": marca,
                "modelo": modelo,
                "año_modelo": anio_modelo,
                "vin": vin,
                "estado": estado or "OPERATIVO",
            },
        )

        if not created:
            if marca:
                veh.marca = marca
            if modelo:
                veh.modelo = modelo
            veh.año_modelo = anio_modelo
            if estado:
                veh.estado = estado
            if vin:
                veh.vin = vin
            veh.save()

        ok += 1

    return ok, errores
